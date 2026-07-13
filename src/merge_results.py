import json
import os
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT = Path(os.getenv("INPUT_DIR", "downloaded"))
OUT = Path(os.getenv("OUT_DIR", "final"))
OUT.mkdir(parents=True, exist_ok=True)


def score_row(r):
    score = 0
    channels = []
    risks = []
    reject = []

    name = str(r.get("名称", ""))
    if "ST" in name.upper() or "退" in name:
        score -= 20
        risks.append("ST/退市风险")
        reject.append("风险标识股票")

    if r.get("遍历状态") == "成功":
        score += 5
    else:
        reject.append("历史数据失败、超时或不足")

    avg = pd.to_numeric(r.get("20日平均成交额"), errors="coerce")
    if pd.notna(avg):
        if avg >= 2e8:
            score += 8
        elif avg >= 8e7:
            score += 5
        elif avg >= 3e7:
            score += 2
        else:
            score -= 8
            risks.append("流动性偏低")
    else:
        risks.append("缺少成交额")

    reversal = 0
    dd = pd.to_numeric(r.get("距5年最高价回撤"), errors="coerce")
    pos = pd.to_numeric(r.get("近250日位置"), errors="coerce")
    tight = pd.to_numeric(r.get("均线粘合度"), errors="coerce")
    ran = pd.to_numeric(r.get("近120日区间"), errors="coerce")
    vr = pd.to_numeric(r.get("量能20比60"), errors="coerce")
    r60 = pd.to_numeric(r.get("近60日收益"), errors="coerce")

    if pd.notna(dd):
        reversal += 8 if dd <= -0.65 else 5 if dd <= -0.45 else 0
    if pd.notna(pos) and pos <= 0.35:
        reversal += 4
    if pd.notna(tight) and tight <= 0.10:
        reversal += 4
    if pd.notna(ran) and ran <= 0.35:
        reversal += 3
    if pd.notna(vr) and vr >= 1.3:
        reversal += 3
    if pd.notna(r60) and 0 <= r60 <= 0.30:
        reversal += 2

    score += reversal
    if reversal >= 13:
        channels.append("深跌筑底")

    trend = 0
    r250 = pd.to_numeric(r.get("近250日收益"), errors="coerce")
    if pd.notna(r60) and 0.05 <= r60 <= 0.40:
        trend += 4
    if pd.notna(r250) and 0.10 <= r250 <= 0.80:
        trend += 4
    if pd.notna(vr) and 1.1 <= vr <= 2.5:
        trend += 3

    score += trend
    if trend >= 9:
        channels.append("趋势景气")

    if pd.notna(r60) and r60 > 0.70:
        score -= 10
        risks.append("近60日涨幅过大")
    if pd.notna(pos) and pos > 0.92:
        score -= 5
        risks.append("接近一年高位")
    if pd.notna(vr) and vr > 4:
        score -= 4
        risks.append("成交过热")

    if not channels:
        reject.append("未触发强技术通道")

    return pd.Series({
        "总评分": round(score, 2),
        "入选通道": "；".join(channels),
        "主要风险": "；".join(dict.fromkeys(risks)),
        "淘汰原因": "；".join(dict.fromkeys(reject)),
    })


def classify(r):
    if r.get("遍历状态") != "成功":
        return "数据异常"
    if r.get("淘汰原因"):
        return "淘汰"
    s = r.get("总评分", -999)
    if s >= 30:
        return "A初筛"
    if s >= 22:
        return "B初筛"
    if s >= 16:
        return "C观察"
    return "淘汰"


def main():
    shard_files = sorted(INPUT.rglob("shard_*.csv"))
    shard_files = [p for p in shard_files if "checkpoint" not in p.name]
    if not shard_files:
        raise RuntimeError("No shard result files found")

    frames = [pd.read_csv(p, dtype={"代码": str}) for p in shard_files]
    scored = pd.concat(frames, ignore_index=True)
    scored["代码"] = scored["代码"].str.zfill(6)
    scored = scored.drop_duplicates("代码")

    scored = scored.join(scored.apply(score_row, axis=1))
    scored["初筛层级"] = scored.apply(classify, axis=1)

    order = {"A初筛": 0, "B初筛": 1, "C观察": 2, "淘汰": 3, "数据异常": 4}
    scored["_order"] = scored["初筛层级"].map(order).fillna(9)
    scored = scored.sort_values(["_order", "总评分"], ascending=[True, False]).drop(columns="_order")

    candidates = scored[scored["初筛层级"].isin(["A初筛", "B初筛", "C观察"])].copy()
    errors = scored[scored["初筛层级"] == "数据异常"].copy()

    total = len(scored)
    success = int(scored["遍历状态"].eq("成功").sum())
    coverage = success / total if total else 0

    coverage_df = pd.DataFrame({
        "指标": ["全市场结果数量", "历史行情成功", "失败/超时/不足", "成功覆盖率",
                 "A初筛", "B初筛", "C观察", "淘汰", "数据异常"],
        "数值": [total, success, total - success, f"{coverage:.2%}",
                 int((scored["初筛层级"] == "A初筛").sum()),
                 int((scored["初筛层级"] == "B初筛").sum()),
                 int((scored["初筛层级"] == "C观察").sum()),
                 int((scored["初筛层级"] == "淘汰").sum()),
                 int((scored["初筛层级"] == "数据异常").sum())],
    })

    readme = pd.DataFrame({"说明": [
        "本文件是全A程序化第一阶段技术与量价扫描，不是最终买入建议。",
        "数据异常股票不能视为淘汰，应针对失败清单补抓。",
        "最终股票池需要进一步核验财务、公告、行业、估值、控制权、解禁减持和拥挤度。",
    ]})

    xlsx = OUT / "A股全市场逐股扫描结果.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="使用说明", index=False)
        coverage_df.to_excel(writer, sheet_name="覆盖率", index=False)
        candidates.to_excel(writer, sheet_name="候选股票池", index=False)
        scored.to_excel(writer, sheet_name="全A逐股结果", index=False)
        errors.to_excel(writer, sheet_name="数据异常", index=False)

    scored.to_csv(OUT / "全A逐股结果.csv", index=False, encoding="utf-8-sig")
    candidates.to_csv(OUT / "候选股票池.csv", index=False, encoding="utf-8-sig")
    errors.to_csv(OUT / "失败股票清单.csv", index=False, encoding="utf-8-sig")

    report = {
        "total": total,
        "success": success,
        "coverage": coverage,
        "levels": scored["初筛层级"].value_counts().to_dict(),
    }
    (OUT / "覆盖率报告.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    wb = load_workbook(xlsx)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in list(col)[:250]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 45))
            ws.column_dimensions[get_column_letter(i)].width = max(10, min(max_len + 2, 45))
    wb.save(xlsx)

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
